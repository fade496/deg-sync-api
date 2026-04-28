#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import math
import re
from collections import defaultdict
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.units import inch
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.platypus import BaseDocTemplate, Frame, PageBreak, PageTemplate, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

CSV_HEADERS = [
    "LINETYPE",
    "WORKDATE",
    "REFWO",
    "LABORCODE",
    "LABORDESC",
    "LABORISNIGHTSHIFT",
    "LABORPAYCODE",
    "LABORHRS",
    "CRAFT+SKILL/PREM",
]

DETAIL_HEADERS = ["Index", "Date", "LEM Name", "Employee", "Project", "Description", "Hours", "WO"]
PAGE_TEMPLATE_ROWS = 61
DETAIL_START_ROW = 8
DETAIL_END_ROW = 57
DETAIL_ROWS_PER_PAGE = DETAIL_END_ROW - DETAIL_START_ROW + 1
DATE_FMT = "%m/%d/%y"


@dataclass
class ProjectInfo:
    code: str
    project_name: str
    billing_type: str
    approver_first: str
    approver_last: str
    approver_email: str
    contract: str
    short_code: str

    @property
    def approver_name(self) -> str:
        return f"{self.approver_first} {self.approver_last}".strip()


@dataclass
class PersonInfo:
    name: str
    craft1: str
    craft2: str
    craft3: str


class LemBuildError(Exception):
    pass


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Generate LEM CSV imports, XLSX approval reports, and native PDFs.")
    p.add_argument("--timesheet", required=True, help="Path to CSV or XLSX timesheet")
    p.add_argument("--airtable-json", required=True, help="Path to normalized Airtable JSON export")
    p.add_argument("--output-dir", required=True, help="Directory for generated files")
    p.add_argument("--template", help="Path to XLSX report template")
    return p.parse_args()


def main() -> int:
    args = parse_args()
    try:
        output_dir = Path(args.output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        errors: list[str] = []
        airtable = load_airtable(Path(args.airtable_json))
        rows = load_timesheet(Path(args.timesheet))
        normalized = normalize_rows(rows, airtable, errors)
        if not normalized:
            raise LemBuildError("No valid rows were produced from the timesheet.")
        write_contract_csvs(normalized, output_dir)
        create_reports(normalized, output_dir, Path(args.template) if args.template else None)
        if errors:
            (output_dir / "errors.txt").write_text("\n".join(errors), encoding="utf-8")
        return 0
    except Exception as exc:  # noqa: BLE001
        Path(args.output_dir).mkdir(parents=True, exist_ok=True)
        (Path(args.output_dir) / "errors.txt").write_text(str(exc), encoding="utf-8")
        raise


def load_airtable(path: Path) -> dict[str, Any]:
    data = json.loads(path.read_text(encoding="utf-8"))
    projects: dict[str, ProjectInfo] = {}
    for item in data.get("projects", []):
        code = str(item.get("Project Code", "")).strip()
        if not code:
            continue
        projects[code] = ProjectInfo(
            code=code,
            project_name=str(item.get("Project", "")).strip(),
            billing_type=str(item.get("Billing Type", "")).strip(),
            approver_first=str(item.get("Approver First Name", "")).strip(),
            approver_last=str(item.get("Approver Last Name", "")).strip(),
            approver_email=str(item.get("Approver Email", "")).strip(),
            contract=str(item.get("Contracts", "")).strip(),
            short_code=str(item.get("Short Code", "")).strip(),
        )
    project_billing = {
        str(item.get("Project Code", "")).strip(): str(item.get("Billing Method", "")).strip()
        for item in data.get("project_billing", [])
        if str(item.get("Project Code", "")).strip()
    }
    people: dict[str, PersonInfo] = {}
    for item in data.get("people", []):
        first = str(item.get("First Name", "")).strip()
        last = str(item.get("Last Name", "")).strip()
        name = f"{first} {last}".strip()
        if not name:
            continue
        people[name] = PersonInfo(
            name=name,
            craft1=str(item.get("Craft Code 1", "")).strip(),
            craft2=str(item.get("Craft Code 2", "")).strip(),
            craft3=str(item.get("Craft Code 3", "")).strip(),
        )
    return {"projects": projects, "project_billing": project_billing, "people": people}


def load_timesheet(path: Path) -> list[dict[str, str]]:
    if path.suffix.lower() == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as f:
            return [dict(row) for row in csv.DictReader(f)]
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v not in (None, "") for v in row):
            continue
        out.append({header[i]: "" if row[i] is None else str(row[i]) for i in range(len(header))})
    return out


def normalize_rows(rows: list[dict[str, str]], airtable: dict[str, Any], errors: list[str]) -> list[dict[str, Any]]:
    out = []
    for raw in rows:
        project_code = raw.get("Project Code", "").strip()
        project = airtable["projects"].get(project_code)
        if not project:
            errors.append(f"Missing project for Project Code: {project_code}")
            continue
        if project.billing_type and project.billing_type.upper() != "LEM":
            continue
        employee_name = build_employee_name(raw)
        person = airtable["people"].get(employee_name)
        if not person:
            errors.append(f"Missing people record for employee: {employee_name}")
            continue
        billing_method = airtable["project_billing"].get(project_code, "Craft 1")
        craft_code = choose_craft_code(billing_method, person)
        work_date = parse_date(raw.get("Date", ""))
        if work_date is None:
            errors.append(f"Invalid date for row: {raw.get('Date', '')}")
            continue
        hours = parse_hours(raw.get("Hours", ""))
        if hours is None:
            errors.append(f"Invalid hours for row: {raw.get('Hours', '')}")
            continue
        description = (raw.get("Notes") or "").strip() or (raw.get("Task") or "").strip() or project.project_name
        wo = extract_wo(raw.get("Task", ""), raw.get("Notes", ""))
        labor_code = extract_id_digits(raw.get("Employee Id", ""))
        report_name = ""
        out.append(
            {
                "contract": project.contract,
                "short_code": project.short_code,
                "project_code": project.code,
                "project_name": project.project_name,
                "approver_first": project.approver_first,
                "approver_last": project.approver_last,
                "approver_name": project.approver_name,
                "approver_email": project.approver_email,
                "employee_name": employee_name,
                "work_date": work_date,
                "work_date_str": work_date.strftime(DATE_FMT),
                "hours": hours,
                "hours_str": format_hours(hours),
                "wo": wo,
                "labor_code": labor_code,
                "craft_code": craft_code,
                "report_name": report_name,
                "description": description,
                "csv_description": employee_name,
            }
        )
    out.sort(key=lambda r: (r["contract"], r["approver_name"], r["work_date"], r["employee_name"]))
    return out


def build_employee_name(raw: dict[str, str]) -> str:
    name = (raw.get("Name") or "").strip()
    if name:
        return name
    return f"{(raw.get('First Name') or '').strip()} {(raw.get('Last Name') or '').strip()}".strip()


def choose_craft_code(billing_method: str, person: PersonInfo) -> str:
    method = billing_method.strip().lower()
    if method == "craft 2":
        return person.craft2 or person.craft1
    if method == "craft 3":
        return person.craft3 or person.craft1
    return person.craft1


def parse_date(value: str) -> datetime | None:
    value = value.strip()
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%m-%d-%Y"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None


def parse_hours(value: str) -> float | None:
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return None


def extract_wo(*values: str) -> str:
    for value in values:
        match = re.search(r"WO-(\d+)", value or "", re.IGNORECASE)
        if match:
            return match.group(1)
    digits = re.search(r"(\d{6,})", " ".join(values))
    return digits.group(1) if digits else ""


def extract_id_digits(value: str) -> str:
    digits = re.findall(r"\d+", value or "")
    return digits[0] if digits else str(value).strip()


def format_hours(value: float) -> str:
    return f"{int(value)}" if float(value).is_integer() else f"{value:.2f}".rstrip("0").rstrip(".")


def contract_base_name(rows: list[dict[str, Any]]) -> str:
    dates = sorted(r["work_date"] for r in rows)
    start = dates[0]
    end = dates[-1]
    short_code = rows[0]["short_code"] or "LEM"
    if start.month == end.month:
        date_part = f"{start:%m.%d}-{end:%d}.{end:%Y}"
    else:
        date_part = f"{start:%m.%d}-{end:%m.%d}.{end:%Y}"
    return f"{date_part}.{short_code}"


def approver_lem_name(first: str, last: str) -> str:
    return f"{first}{last}".replace(" ", "")


def write_contract_csvs(rows: list[dict[str, Any]], output_dir: Path) -> None:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[row["contract"]].append(row)
    for contract, group in grouped.items():
        base_name = contract_base_name(group)
        csv_path = output_dir / f"{base_name}.csv"
        with csv_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["CNRLEMLINE", base_name, contract, "O"])
            writer.writerow(CSV_HEADERS)
            for row in group:
                writer.writerow([
                    "L",
                    row["work_date_str"],
                    row["wo"],
                    row["labor_code"],
                    row["employee_name"],
                    "0",
                    "801",
                    row["hours_str"],
                    row["craft_code"],
                ])


def create_reports(rows: list[dict[str, Any]], output_dir: Path, template_path: Path | None) -> None:
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[(row["contract"], row["approver_name"])].append(row)
    for (_, approver_name), group in grouped.items():
        contract_name = contract_base_name(group)
        safe_name = f"{contract_name}.{group[0]['approver_first']}{group[0]['approver_last']}".replace(" ", "")
        xlsx_path = output_dir / f"{safe_name}.xlsx"
        pdf_path = output_dir / f"{safe_name}.pdf"
        for row in group:
            row["report_name"] = safe_name
        build_report_workbook(group, xlsx_path, template_path)
        build_report_pdf(group, pdf_path)


def build_report_workbook(rows: list[dict[str, Any]], output_path: Path, template_path: Path | None) -> None:
    if template_path and template_path.exists():
        wb = load_workbook(template_path)
        source_ws = wb[wb.sheetnames[0]]
    else:
        wb = Workbook()
        source_ws = wb.active
        source_ws.title = "LEM Report"
        seed_default_template(source_ws)
    ws = source_ws
    pages = max(1, math.ceil(len(rows) / DETAIL_ROWS_PER_PAGE))
    if pages > 1:
        expand_template_pages(ws, pages)
    fill_report_headers(ws, rows)
    fill_report_rows(ws, rows)
    apply_page_setup(ws, pages)
    wb.save(output_path)


def seed_default_template(ws) -> None:
    ws.title = "LEM Report"
    ws["A1"] = "DEG engineering"
    ws["A2"] = "Client:"
    ws["F2"] = "Report Name:"
    ws["A3"] = "Project:"
    ws["F3"] = "Report Approver:"
    ws["A4"] = "Date Generated:"
    ws["F4"] = "Total Hours:"
    for i, header in enumerate(DETAIL_HEADERS, start=1):
        ws.cell(7, i, header)
    ws["A58"] = "Total Hours"
    ws["A61"] = "DEG engineering Inc. • Confidential Report"
    ws["E61"] = "Approved By:"


def expand_template_pages(ws, pages: int) -> None:
    if pages <= 1:
        return
    for page_index in range(2, pages + 1):
        start_target = (page_index - 1) * PAGE_TEMPLATE_ROWS + 1
        insert_at = start_target
        ws.insert_rows(insert_at, PAGE_TEMPLATE_ROWS)
        copy_row_block(ws, 1, PAGE_TEMPLATE_ROWS, insert_at)


def copy_row_block(ws, source_start: int, source_count: int, target_start: int) -> None:
    for row_offset in range(source_count):
        src_row = source_start + row_offset
        dst_row = target_start + row_offset
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
        for col in range(1, ws.max_column + 1):
            src = ws.cell(src_row, col)
            dst = ws.cell(dst_row, col)
            if src.data_type == 'f':
                dst.value = translate_formula(src.value, dst_row - src_row)
            else:
                dst.value = src.value
            if src.has_style:
                dst._style = copy(src._style)
            if src.number_format:
                dst.number_format = src.number_format
            if src.font:
                dst.font = copy(src.font)
            if src.fill:
                dst.fill = copy(src.fill)
            if src.border:
                dst.border = copy(src.border)
            if src.alignment:
                dst.alignment = copy(src.alignment)
            if src.protection:
                dst.protection = copy(src.protection)
    copy_merged_ranges(ws, source_start, source_count, target_start)


def copy_merged_ranges(ws, source_start: int, source_count: int, target_start: int) -> None:
    source_end = source_start + source_count - 1
    existing = list(ws.merged_cells.ranges)
    for merged in existing:
        if merged.min_row >= source_start and merged.max_row <= source_end:
            row_offset = target_start - source_start
            new_min_row = merged.min_row + row_offset
            new_max_row = merged.max_row + row_offset
            ws.merge_cells(
                start_row=new_min_row,
                start_column=merged.min_col,
                end_row=new_max_row,
                end_column=merged.max_col,
            )


def translate_formula(formula: str, row_delta: int) -> str:
    def repl(match: re.Match[str]) -> str:
        col = match.group(1)
        row = int(match.group(2))
        return f"{col}{row + row_delta}"
    return re.sub(r"([A-Z]+)(\d+)", repl, formula)


def fill_report_headers(ws, rows: list[dict[str, Any]]) -> None:
    total_hours = sum(r["hours"] for r in rows)
    project_names = sorted({r["project_name"] for r in rows})
    project_label = project_names[0] if len(project_names) == 1 else "Multiple Projects"
    report_name = rows[0]["report_name"]
    for page_idx in range(max(1, math.ceil(len(rows) / DETAIL_ROWS_PER_PAGE))):
        offset = page_idx * PAGE_TEMPLATE_ROWS
        ws[f"B{2 + offset}"] = rows[0]["contract"]
        ws[f"G{2 + offset}"] = report_name
        ws[f"B{3 + offset}"] = project_label
        ws[f"G{3 + offset}"] = rows[0]["approver_name"]
        ws[f"B{4 + offset}"] = datetime.now().strftime("%m/%d/%Y")
        ws[f"G{4 + offset}"] = total_hours
        sig_row = 61 + offset
        ws[f"G{sig_row}"] = rows[0]["approver_name"]
        ws[f"G{sig_row}"].alignment = Alignment(horizontal="left")


def fill_report_rows(ws, rows: list[dict[str, Any]]) -> None:
    for idx, row in enumerate(rows, start=1):
        page = (idx - 1) // DETAIL_ROWS_PER_PAGE
        slot = (idx - 1) % DETAIL_ROWS_PER_PAGE
        excel_row = page * PAGE_TEMPLATE_ROWS + DETAIL_START_ROW + slot
        ws[f"A{excel_row}"] = idx
        ws[f"B{excel_row}"] = row["work_date"].strftime("%m/%d/%Y")
        ws[f"C{excel_row}"] = row["report_name"]
        ws[f"D{excel_row}"] = row["employee_name"]
        ws[f"E{excel_row}"] = row["project_name"]
        ws[f"F{excel_row}"] = row["description"]
        ws[f"G{excel_row}"] = row["hours"]
        ws[f"H{excel_row}"] = row["wo"]
    pages = max(1, math.ceil(len(rows) / DETAIL_ROWS_PER_PAGE))
    total_hours = sum(r["hours"] for r in rows)
    for page in range(pages):
        total_row = 58 + page * PAGE_TEMPLATE_ROWS
        ws[f"G{total_row}"] = total_hours


def apply_page_setup(ws, pages: int) -> None:
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.sheet_view.zoomScale = 100
    page_ranges = []
    for page in range(pages):
        start = 1 + page * PAGE_TEMPLATE_ROWS
        end = PAGE_TEMPLATE_ROWS + page * PAGE_TEMPLATE_ROWS
        page_ranges.append(f"$A${start}:$H${end}")
    ws.print_area = ",".join(page_ranges)
    ws.page_margins.left = 0.35
    ws.page_margins.right = 0.35
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35


def build_report_pdf(rows: list[dict[str, Any]], output_path: Path) -> None:
    page_width, page_height = landscape(letter)
    left = 0.35 * inch
    right = 0.35 * inch
    top = 0.35 * inch
    bottom = 0.35 * inch
    usable_width = page_width - left - right
    usable_height = page_height - top - bottom
    header_h = 1.10 * inch
    signature_h = 0.65 * inch
    table_h = usable_height - header_h - signature_h - 0.08 * inch
    doc = BaseDocTemplate(str(output_path), pagesize=landscape(letter), leftMargin=left, rightMargin=right, topMargin=top, bottomMargin=bottom)
    frame = Frame(left, bottom + signature_h, usable_width, table_h, id="detail")

    def on_page(canvas, _doc):
        draw_pdf_page_decor(canvas, rows, left, bottom, usable_width, page_height - top)

    doc.addPageTemplates([PageTemplate(id="lem", frames=[frame], onPage=on_page)])

    styles = getSampleStyleSheet()
    body = ParagraphStyle("body", parent=styles["BodyText"], fontName="Helvetica", fontSize=8, leading=10)
    data = [DETAIL_HEADERS]
    for idx, row in enumerate(rows, start=1):
        data.append([
            str(idx),
            row["work_date"].strftime("%m/%d/%Y"),
            row["report_name"],
            row["employee_name"],
            row["project_name"],
            Paragraph(escape_pdf(row["description"]), body),
            row["hours_str"],
            row["wo"],
        ])
    col_widths = fit_column_widths(data, usable_width, styles)
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9D9D9")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("LEADING", (0, 0), (-1, -1), 10),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (1, 1), (2, -1), "LEFT"),
        ("ALIGN", (6, 1), (7, -1), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    doc.build([table])


def draw_pdf_page_decor(canvas, rows: list[dict[str, Any]], left: float, bottom: float, usable_width: float, top_y: float) -> None:
    total_hours = sum(r["hours"] for r in rows)
    project_names = sorted({r["project_name"] for r in rows})
    project_label = project_names[0] if len(project_names) == 1 else "Multiple Projects"
    report_name = rows[0]["report_name"]
    client = rows[0]["contract"]
    approver = rows[0]["approver_name"]

    canvas.setFont("Helvetica-Bold", 16)
    canvas.drawString(left, top_y + 2, "DEG engineering")

    box_top = top_y - 14
    row_h = 18
    label_w = 86
    value_w = usable_width / 2 - label_w - 6
    x1 = left
    x2 = left + usable_width / 2
    labels1 = [("Client:", client), ("Project:", project_label), ("Date Generated:", datetime.now().strftime("%m/%d/%Y"))]
    labels2 = [("Report Name:", report_name), ("Report Approver:", approver), ("Total Hours:", format_hours(total_hours))]
    for i, (label, value) in enumerate(labels1):
        y = box_top - i * row_h
        draw_header_cell(canvas, x1, y, label_w, row_h, label, bold=True)
        draw_header_cell(canvas, x1 + label_w, y, value_w, row_h, value)
    for i, (label, value) in enumerate(labels2):
        y = box_top - i * row_h
        draw_header_cell(canvas, x2, y, label_w, row_h, label, bold=True)
        draw_header_cell(canvas, x2 + label_w, y, value_w, row_h, value)

    footer_y = bottom + 16
    canvas.setFont("Helvetica-Bold", 9)
    canvas.drawString(left + 18, footer_y + 18, "Signature:")
    canvas.drawString(left + usable_width * 0.72, footer_y + 18, "Date:")


def draw_header_cell(canvas, x: float, y: float, w: float, h: float, text: str, bold: bool = False) -> None:
    canvas.setFont("Helvetica-Bold" if bold else "Helvetica", 8)
    canvas.drawString(x + 4, y - 12, str(text))


def fit_column_widths(data: list[list[Any]], total_width: float, styles) -> list[float]:
    raw_widths = []
    sample_style = styles["BodyText"]
    for col_idx in range(len(DETAIL_HEADERS)):
        max_width = stringWidth(str(DETAIL_HEADERS[col_idx]), "Helvetica-Bold", 8) + 12
        for row in data[1:]:
            value = row[col_idx]
            if isinstance(value, Paragraph):
                width = 180
            else:
                width = stringWidth(str(value), "Helvetica", 8) + 10
            max_width = max(max_width, width)
        raw_widths.append(max_width)
    scale = total_width / sum(raw_widths)
    widths = [w * scale for w in raw_widths]
    min_widths = [28, 54, 75, 90, 110, 170, 42, 52]
    widths = [max(widths[i], min_widths[i]) for i in range(len(widths))]
    total = sum(widths)
    if total > total_width:
        overflow = total - total_width
        widths[5] = max(min_widths[5], widths[5] - overflow)
    return widths


def escape_pdf(value: str) -> str:
    return (
        str(value)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


if __name__ == "__main__":
    raise SystemExit(main())
